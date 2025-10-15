"""
Unicommerce Pivot Table Creator
===============================

This module creates pivot tables for Unicommerce export data by:
1. Merging all facility data into a master sheet
2. Creating a formatted pivot table for analysis
3. Integrating with the existing export workflow

Features:
- Master data sheet creation from all facilities
- Channel-wise analysis for complete orders only
- Automatic integration with export job completion
- Email notification with pivot table Excel file
"""

import frappe
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime
from typing import Dict, List, Optional

# =============================================================================
# CONFIGURATION
# =============================================================================

# Styling configuration
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
DATA_FONT = Font(size=10)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _safe_convert_to_float(value) -> float:
    """Safely convert value to float"""
    try:
        if value is None or pd.isna(value) or value == '':
            return 0.0
        return float(str(value).strip())
    except:
        return 0.0

def _clean_facility_name(facility_name: str) -> str:
    """Clean facility name for display"""
    return facility_name.replace('_', ' ').replace('/', ' ').title()

def _extract_month_from_date(date_str: str) -> str:
    """Extract month from date string"""
    try:
        if pd.isna(date_str) or date_str == '':
            return 'Unknown'
        date_obj = pd.to_datetime(date_str)
        return date_obj.strftime('%Y-%m')
    except:
        return 'Unknown'

def _standardize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize column names for consistency"""
    column_mapping = {
        'Sale Order Code': 'Channel Order ID',
        'Display Order Code': 'Channel Order ID',
        'Item SKU Code': 'Item SKU Code',
        'Item Name': 'Item Name',
        'Selling Price': 'Selling Price',
        'Total Price': 'Total Price',
        'Quantity': 'Quantity',
        'Channel Name': 'Channel',
        'Sale Order Status': 'Status',
        'Billing Address Name': 'Customer Name',
        'Billing Address City': 'City',
        'Billing Address State': 'State',
        'Billing Address Pin Code': 'Pin Code',
        'Display Order Date Time': 'Order Date',
        'Created': 'Created Date'
    }
    
    # Rename columns that exist
    for old_name, new_name in column_mapping.items():
        if old_name in df.columns:
            df = df.rename(columns={old_name: new_name})
    
    return df

def _add_calculated_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Add calculated columns for analysis"""
    # Ensure numeric columns are properly converted
    numeric_columns = ['Selling Price', 'Total Price', 'Quantity', 'Subtotal', 'Discount']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].apply(_safe_convert_to_float)
    
    # Add revenue per order
    if 'Total Price' in df.columns:
        df['Revenue'] = df['Total Price']
    else:
        df['Revenue'] = 0.0
    
    # Add Net Taxable Value column: (Subtotal - Discount)
    if 'Subtotal' in df.columns and 'Discount' in df.columns:
        df['Net Taxable Value'] = df['Subtotal'] - df['Discount']
        print(f"[MASTER DATA] Added Net Taxable Value column: (Subtotal - Discount)")
    else:
        df['Net Taxable Value'] = 0.0
        print(f"[MASTER DATA] ⚠️ Subtotal or Discount column not found, Net Taxable Value set to 0")
    
    # Reorder columns to place Net Taxable Value right after Discount
    df = _reorder_columns(df)
    
    return df

def _reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Reorder columns to place Net Taxable Value right after Discount"""
    try:
        # Get current column order
        current_columns = list(df.columns)
        
        # Find the position of Discount column
        discount_index = None
        for i, col in enumerate(current_columns):
            if col == 'Discount':
                discount_index = i
                break
        
        if discount_index is not None:
            # Create new column order
            new_columns = []
            
            # Add columns before Discount
            for i in range(discount_index + 1):
                new_columns.append(current_columns[i])
            
            # Add Net Taxable Value right after Discount
            if 'Net Taxable Value' in current_columns:
                new_columns.append('Net Taxable Value')
            
            # Add remaining columns (excluding Net Taxable Value if it was already added)
            for i in range(discount_index + 1, len(current_columns)):
                if current_columns[i] != 'Net Taxable Value':
                    new_columns.append(current_columns[i])
            
            # Reorder the dataframe
            df = df[new_columns]
            print(f"[MASTER DATA] Reordered columns: Net Taxable Value placed after Discount")
        else:
            print(f"[MASTER DATA] ⚠️ Discount column not found, keeping original column order")
        
        return df
        
    except Exception as e:
        print(f"[MASTER DATA] ⚠️ Error reordering columns: {str(e)}")
        return df

def _reorder_workbook_sheets(workbook):
    """Reorder workbook sheets: Pivot Table first, Master Data second, KFL_ECOM third, then others"""
    try:
        # Get current sheet names
        current_sheets = workbook.sheetnames
        print(f"[SHEET REORDER] Current sheets: {current_sheets}")
        
        # Define the desired order
        desired_order = []
        
        # 1. Pivot Table first
        if 'Pivot Table' in current_sheets:
            desired_order.append('Pivot Table')
        
        # 2. Master Data second
        if 'Master Data' in current_sheets:
            desired_order.append('Master Data')
        
        # 3. KFL_ECOM_WAREHOUSE third
        kfl_sheet = None
        for sheet_name in current_sheets:
            if 'KFL_ECOM' in sheet_name.upper() or 'KFL_ECOM_WAREHOUSE' in sheet_name.upper():
                kfl_sheet = sheet_name
                break
        
        if kfl_sheet:
            desired_order.append(kfl_sheet)
        
        # 4. Add remaining facility sheets (excluding the ones already added)
        for sheet_name in current_sheets:
            if sheet_name not in desired_order:
                desired_order.append(sheet_name)
        
        print(f"[SHEET REORDER] Desired order: {desired_order}")
        
        # Reorder sheets by moving them to the correct positions
        reordered_sheets = []
        for sheet_name in desired_order:
            if sheet_name in workbook.sheetnames:
                reordered_sheets.append(workbook[sheet_name])
        
        # Reorder the workbook sheets
        workbook._sheets = reordered_sheets
        
        print(f"[SHEET REORDER] ✅ Sheets reordered successfully")
        
    except Exception as e:
        print(f"[SHEET REORDER] ⚠️ Error reordering sheets: {str(e)}")

def _style_master_data_sheet(worksheet, df: pd.DataFrame):
    """Apply styling to master data sheet"""
    # Style header row
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    
    # Style data rows
    for row_num in range(2, len(df) + 2):
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = DATA_FONT
            cell.border = BORDER
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

# =============================================================================
# MASTER DATA CREATION
# =============================================================================

@frappe.whitelist()
def create_master_data_sheet(excel_file_path: str) -> Dict:
    """
    Create a master data sheet by merging all facility data
    
    Args:
        excel_file_path: Path to the Excel file with facility sheets
        
    Returns:
        Dictionary with results
    """
    try:
        print(f"[MASTER DATA] Creating master data sheet from: {excel_file_path}")
        
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file_path)
        print(f"[MASTER DATA] Workbook loaded with sheets: {workbook.sheetnames}")
        
        # Collect all data from facility sheets
        all_data = []
        facility_stats = {}
        
        for sheet_name in workbook.sheetnames:
            # Skip summary sheets
            if sheet_name.lower() in ['validation summary', 'master data', 'pivot table']:
                continue
                
            try:
                print(f"[MASTER DATA] Processing sheet: {sheet_name}")
                
                # Read sheet data
                df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
                
                if len(df) > 0:
                    # Add facility column
                    df['Facility'] = _clean_facility_name(sheet_name)
                    
                    # Add month column for trend analysis
                    if 'Display Order Date Time' in df.columns:
                        df['Month'] = df['Display Order Date Time'].apply(_extract_month_from_date)
                    elif 'Created' in df.columns:
                        df['Month'] = df['Created'].apply(_extract_month_from_date)
                    else:
                        df['Month'] = 'Unknown'
                    
                    # Add order count column (each row represents one order)
                    df['Order Count'] = 1
                    
                    # Standardize column names
                    df = _standardize_column_names(df)
                    
                    # Store facility statistics
                    facility_stats[sheet_name] = {
                        'rows': len(df),
                        'facility_name': _clean_facility_name(sheet_name),
                        'has_data': len(df) > 0
                    }
                    
                    all_data.append(df)
                    print(f"[MASTER DATA] Added {len(df)} rows from {sheet_name}")
                
            except Exception as e:
                print(f"[MASTER DATA] Error processing sheet {sheet_name}: {str(e)}")
                continue
        
        if not all_data:
            return {
                'status': 'error',
                'message': 'No data found in facility sheets'
            }
        
        # Merge all data
        master_df = pd.concat(all_data, ignore_index=True, sort=False)
        
        # Fill missing values
        master_df = master_df.fillna('')
        
        # Add calculated columns
        master_df = _add_calculated_columns(master_df)
        
        print(f"[MASTER DATA] Created master data with {len(master_df)} total rows")
        
        # Create master data sheet (will be reordered later)
        if 'Master Data' in workbook.sheetnames:
            workbook.remove(workbook['Master Data'])
        
        master_sheet = workbook.create_sheet('Master Data')
        
        # Write data to sheet
        for r in dataframe_to_rows(master_df, index=False, header=True):
            master_sheet.append(r)
        
        # Style the master data sheet
        _style_master_data_sheet(master_sheet, master_df)
        
        # Save the workbook
        workbook.save(excel_file_path)
        workbook.close()
        
        return {
            'status': 'success',
            'message': f'Master data sheet created with {len(master_df)} rows',
            'total_rows': len(master_df),
            'facilities_processed': len(facility_stats),
            'facility_stats': facility_stats,
            'columns': list(master_df.columns)
        }
        
    except Exception as e:
        error_msg = f"Error creating master data sheet: {str(e)}"
        print(f"[MASTER DATA] ERROR: {error_msg}")
        return {
            'status': 'error',
            'error': error_msg
        }

# =============================================================================
# PIVOT TABLE CREATION
# =============================================================================

@frappe.whitelist()
def create_pivot_tables(excel_file_path: str) -> Dict:
    """
    Create pivot table from master data
    
    Args:
        excel_file_path: Path to the Excel file with master data
        
    Returns:
        Dictionary with results
    """
    try:
        print(f"[PIVOT TABLES] Creating pivot table from: {excel_file_path}")
        
        # Load master data
        master_df = pd.read_excel(excel_file_path, sheet_name='Master Data')
        
        if len(master_df) == 0:
            return {
                'status': 'error',
                'message': 'No data found in master data sheet'
            }
        
        print(f"[PIVOT TABLES] Creating pivot table")
        
        # Load workbook
        workbook = openpyxl.load_workbook(excel_file_path)
        
        # Remove existing pivot table sheet if it exists
        if 'Pivot Table' in workbook.sheetnames:
            workbook.remove(workbook['Pivot Table'])
        
        # Create Excel pivot table
        pivot_created = _create_excel_pivot_table(workbook, master_df)
        
        created_tables = []
        if pivot_created:
            created_tables.append({
                'name': 'Custom Business Analysis',
                'type': 'Excel Pivot Table',
                'description': 'Channel-wise analysis for complete orders'
            })
            print(f"[PIVOT TABLES] ✅ Created Excel Pivot Table")
        else:
            print(f"[PIVOT TABLES] ❌ Failed to create Excel Pivot Table")
        
        # Reorder sheets: Pivot Table first, Master Data second, KFL_ECOM third, then others
        _reorder_workbook_sheets(workbook)
        
        # Save workbook
        workbook.save(excel_file_path)
        workbook.close()
        
        return {
            'status': 'success',
            'message': f'Created {len(created_tables)} pivot tables',
            'created_tables': created_tables
        }
        
    except Exception as e:
        error_msg = f"Error creating pivot tables: {str(e)}"
        print(f"[PIVOT TABLES] ERROR: {error_msg}")
        return {
            'status': 'error',
            'error': error_msg
        }

def _create_excel_pivot_table(workbook, master_df: pd.DataFrame):
    """Create a formatted table that looks like a pivot table"""
    try:
        print(f"[EXCEL PIVOT] Creating formatted pivot table")
        
        # Filter to only COMPLETE orders
        if 'Status' in master_df.columns:
            complete_df = master_df[master_df['Status'].str.upper() == 'COMPLETE']
            print(f"[EXCEL PIVOT] Filtered to complete orders: {len(complete_df)} rows")
        else:
            print(f"[EXCEL PIVOT] No Status column found, using all data")
            complete_df = master_df
        
        if len(complete_df) == 0:
            print(f"[EXCEL PIVOT] No complete orders found")
            return False
        
        # Create pivot data using pandas
        channel_col = 'Channel' if 'Channel' in complete_df.columns else 'Channel Name'
        if channel_col not in complete_df.columns:
            print(f"[EXCEL PIVOT] No Channel column found")
            return False
        
        # Define the value columns we want to sum (removed Selling Price, added Net Taxable Value)
        # Order: Subtotal, Discount, Net Taxable Value, then other columns
        value_columns = ['Subtotal', 'Discount', 'Net Taxable Value', 'CGST', 'SGST', 'IGST', 'UTGST', 'Shipping Charges']
        available_columns = [col for col in value_columns if col in complete_df.columns]
        
        if not available_columns:
            print(f"[EXCEL PIVOT] No value columns found")
            return False
        
        # Create the pivot table
        pivot_data = complete_df.groupby([channel_col]).agg({col: 'sum' for col in available_columns}).reset_index()
        
        # Sort by Net Taxable Value (descending)
        if 'Net Taxable Value' in pivot_data.columns:
            pivot_data = pivot_data.sort_values('Net Taxable Value', ascending=False)
        
        # Create a new sheet for the pivot table (at position 0 - first sheet)
        pivot_sheet = workbook.create_sheet('Pivot Table', 0)
        
        # Add title
        title_cell = pivot_sheet.cell(row=1, column=1)
        title_cell.value = "Custom Business Analysis - Complete Orders Only"
        title_cell.font = Font(bold=True, size=14, color="366092")
        
        # Add description
        desc_cell = pivot_sheet.cell(row=2, column=1)
        desc_cell.value = "Channel-wise analysis with filters: Sale Order Status = COMPLETE"
        desc_cell.font = Font(italic=True, size=10)
        
        # Write the pivot table data starting from row 4
        start_row = 4
        
        # Write headers
        for col_num, column_title in enumerate(pivot_data.columns, 1):
            cell = pivot_sheet.cell(row=start_row, column=col_num)
            cell.value = column_title
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER
        
        # Write data
        for row_idx, row_data in enumerate(pivot_data.values, start_row + 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = pivot_sheet.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                cell.font = DATA_FONT
                cell.border = BORDER
                if col_idx > 1:  # Right-align numeric columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:  # Left-align first column (Channel)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for column in pivot_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            pivot_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create second pivot table below the first one
        _create_second_pivot_table(pivot_sheet, complete_df, len(pivot_data) + 6)
        
        print(f"[EXCEL PIVOT] Created formatted pivot table with {len(pivot_data)} channels")
        return True
        
    except Exception as e:
        print(f"[EXCEL PIVOT] Error creating formatted pivot table: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def _create_second_pivot_table(worksheet, master_df: pd.DataFrame, start_row: int):
    """Create a second pivot table below the first one - Channel-wise analysis for RETURNED and RETURN_EXPECTED orders"""
    try:
        print(f"[SECOND PIVOT] Creating second pivot table starting at row {start_row}")
        
        # Add spacing
        start_row += 2
        
        # Add title for second pivot table
        title_cell = worksheet.cell(row=start_row, column=1)
        title_cell.value = "Channel-wise Analysis - Returned Orders Only"
        title_cell.font = Font(bold=True, size=14, color="366092")
        
        # Add description
        desc_cell = worksheet.cell(row=start_row + 1, column=1)
        desc_cell.value = "Channel-wise analysis with filters: Shipping Package Status Code = RETURNED, RETURN_EXPECTED"
        desc_cell.font = Font(italic=True, size=10)
        
        # Start data from row start_row + 3
        data_start_row = start_row + 3
        
        # Filter data by Shipping Package Status Code
        if 'Shipping Package Status Code' in master_df.columns:
            # Filter for RETURNED and RETURN_EXPECTED orders
            returned_df = master_df[
                master_df['Shipping Package Status Code'].isin(['RETURNED', 'RETURN_EXPECTED'])
            ]
            print(f"[SECOND PIVOT] Filtered to returned orders: {len(returned_df)} rows")
        else:
            print(f"[SECOND PIVOT] No 'Shipping Package Status Code' column found, using all data")
            returned_df = master_df
        
        if len(returned_df) == 0:
            print(f"[SECOND PIVOT] No returned orders found")
            # Add a message indicating no data
            no_data_cell = worksheet.cell(row=data_start_row, column=1)
            no_data_cell.value = "No returned orders found in the data"
            no_data_cell.font = Font(italic=True, size=12, color="FF0000")
            return True
        
        # Create the same pivot table as the first one but with filtered data
        channel_col = 'Channel' if 'Channel' in returned_df.columns else 'Channel Name'
        if channel_col not in returned_df.columns:
            print(f"[SECOND PIVOT] No Channel column found")
            # Add a message indicating no channel data
            no_channel_cell = worksheet.cell(row=data_start_row, column=1)
            no_channel_cell.value = "No Channel column found in the data"
            no_channel_cell.font = Font(italic=True, size=12, color="FF0000")
            return True
        
        # Define the value columns we want to sum (same as first pivot table - removed Selling Price, added Net Taxable Value)
        # Order: Subtotal, Discount, Net Taxable Value, then other columns
        value_columns = ['Subtotal', 'Discount', 'Net Taxable Value', 'CGST', 'SGST', 'IGST', 'UTGST', 'Shipping Charges']
        available_columns = [col for col in value_columns if col in returned_df.columns]
        
        if not available_columns:
            print(f"[SECOND PIVOT] No value columns found")
            # Add a message indicating no value columns
            no_value_cell = worksheet.cell(row=data_start_row, column=1)
            no_value_cell.value = "No value columns found in the data"
            no_value_cell.font = Font(italic=True, size=12, color="FF0000")
            return True
        
        # Create the pivot table (same structure as first pivot table)
        pivot_data = returned_df.groupby([channel_col]).agg({col: 'sum' for col in available_columns}).reset_index()
        
        # Sort by Net Taxable Value (descending)
        if 'Net Taxable Value' in pivot_data.columns:
            pivot_data = pivot_data.sort_values('Net Taxable Value', ascending=False)
        
        # Write headers
        for col_num, column_title in enumerate(pivot_data.columns, 1):
            cell = worksheet.cell(row=data_start_row, column=col_num)
            cell.value = column_title
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER
        
        # Write data
        for row_idx, row_data in enumerate(pivot_data.values, data_start_row + 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                cell.font = DATA_FONT
                cell.border = BORDER
                if col_idx > 1:  # Right-align numeric columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:  # Left-align first column (Channel)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        print(f"[SECOND PIVOT] Created channel-wise pivot table for returned orders with {len(pivot_data)} channels")
        
        return True
        
    except Exception as e:
        print(f"[SECOND PIVOT] Error creating second pivot table: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

# =============================================================================
# COMPLETE WORKFLOW
# =============================================================================

@frappe.whitelist()
def create_complete_pivot_analysis(excel_file_path: str, email_notification: bool = True) -> Dict:
    """
    Complete workflow: Create master data sheet and pivot table
    
    Args:
        excel_file_path: Path to the Excel file with facility sheets
        email_notification: Whether to send email notification
        
    Returns:
        Dictionary with complete results
    """
    try:
        print(f"[COMPLETE ANALYSIS] Start: {excel_file_path}")
        
        # Check if file exists
        if not os.path.exists(excel_file_path):
            error_msg = f"Excel file not found: {excel_file_path}"
            print(f"[COMPLETE ANALYSIS] ERROR: {error_msg}")
            return {
                'status': 'error',
                'error': error_msg
            }
        
        # Step 1: Create master data sheet
        print("[COMPLETE ANALYSIS] Master data...")
        master_result = create_master_data_sheet(excel_file_path)
        # minimal log
        
        if master_result['status'] != 'success':
            error_msg = f"Failed to create master data sheet: {master_result.get('error', 'Unknown error')}"
            print(f"[COMPLETE ANALYSIS] ERROR: {error_msg}")
            return {
                'status': 'error',
                'message': 'Failed to create master data sheet',
                'error': master_result.get('error', 'Unknown error'),
                'master_data_result': master_result
            }
        
        # Step 2: Create pivot table
        print("[COMPLETE ANALYSIS] Pivot...")
        pivot_result = create_pivot_tables(excel_file_path)
        # minimal log
        
        if pivot_result['status'] != 'success':
            error_msg = f"Failed to create pivot table: {pivot_result.get('error', 'Unknown error')}"
            print(f"[COMPLETE ANALYSIS] ERROR: {error_msg}")
            return {
                'status': 'error',
                'message': 'Failed to create pivot table',
                'error': pivot_result.get('error', 'Unknown error'),
                'master_data': master_result,
                'pivot_table_result': pivot_result
            }
        
        # Step 3: Send email notification if requested
        if email_notification:
            print("[COMPLETE ANALYSIS] Email...")
            try:
                _send_pivot_analysis_email(excel_file_path, master_result, pivot_result)
            except Exception as e:
                print(f"[COMPLETE ANALYSIS] Warning: Email notification failed: {str(e)}")
        
        print("[COMPLETE ANALYSIS] ✅ Done")
        return {
            'status': 'success',
            'message': 'Complete pivot analysis created successfully',
            'master_data': master_result,
            'pivot_table': pivot_result,
            'excel_file': excel_file_path,
            'email_sent': email_notification
        }
        
    except Exception as e:
        error_msg = f"Error in complete pivot analysis: {str(e)}"
        print(f"[COMPLETE ANALYSIS] ERROR: {error_msg}")
        import traceback
        traceback.print_exc()
        return {
            'status': 'error',
            'error': error_msg
        }

def _send_pivot_analysis_email(excel_file_path: str, master_result: Dict, pivot_result: Dict):
    """Send email notification with pivot analysis results"""
    try:
        # Get file info
        file_size = os.path.getsize(excel_file_path) / (1024 * 1024)  # MB
        file_name = os.path.basename(excel_file_path)
        
        # Prepare email content
        subject = f"Unicommerce Pivot Analysis Complete - {file_name}"
        
        message = f"""
        <h2>Unicommerce Pivot Analysis Complete</h2>
        
        <h3>📊 Analysis Summary</h3>
        <ul>
            <li><strong>Total Records:</strong> {master_result.get('total_rows', 0):,}</li>
            <li><strong>Facilities Processed:</strong> {master_result.get('facilities_processed', 0)}</li>
            <li><strong>Pivot Tables Created:</strong> {len(pivot_result.get('created_tables', []))}</li>
            <li><strong>File Size:</strong> {file_size:.2f} MB</li>
        </ul>
        
        <h3>📈 Pivot Table Created</h3>
        <ul>
        """
        
        for table in pivot_result.get('created_tables', []):
            message += f"<li><strong>{table['name']}:</strong> {table['description']}</li>"
        
        message += f"""
        </ul>
        
        <h3>📁 File Details</h3>
        <ul>
            <li><strong>File Name:</strong> {file_name}</li>
            <li><strong>Location:</strong> {excel_file_path}</li>
            <li><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</li>
        </ul>
        
        <p>The Excel file contains:</p>
        <ul>
            <li><strong>Master Data Sheet:</strong> Combined data from all facilities</li>
            <li><strong>Pivot Table Sheet:</strong> Channel-wise analysis for complete orders</li>
            <li><strong>Individual Facility Sheets:</strong> Original facility data</li>
        </ul>
        
        <p>Best regards,<br>Unicommerce Automation System</p>
        """
        
        # Send email
        frappe.sendmail(
            recipients=['yogesha@khanalfoods.com', 'harsha@khanalfoods.com'],
            subject=subject,
            message=message,
            attachments=[{
                'fname': file_name,
                'fcontent': open(excel_file_path, 'rb').read()
            }]
        )
        
        print("[EMAIL] Pivot analysis email sent successfully")
        
    except Exception as e:
        print(f"[EMAIL] Error sending email: {str(e)}")
        raise

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

@frappe.whitelist()
def create_pivot_analysis_for_existing_file(excel_file_path: str) -> Dict:
    """
    Create pivot analysis for an existing Excel file
    
    This function can be called to add pivot analysis to an existing Excel file
    that already has facility data sheets.
    """
    try:
        print(f"[EXISTING FILE] Creating pivot analysis for existing file: {excel_file_path}")
        
        # Check if file exists
        if not os.path.exists(excel_file_path):
            return {
                'status': 'error',
                'message': f'Excel file not found: {excel_file_path}'
            }
        
        # Create complete pivot analysis
        result = create_complete_pivot_analysis(excel_file_path, email_notification=True)
        
        return result
        
    except Exception as e:
        error_msg = f"Error creating pivot analysis for existing file: {str(e)}"
        print(f"[EXISTING FILE] ERROR: {error_msg}")
        return {
            'status': 'error',
            'error': error_msg
        }