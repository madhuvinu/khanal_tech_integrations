from .unicommerceFile.unicommerce import AuthenticateUniware
import requests
import json
import frappe
from frappe.utils import today, add_months, get_first_day, get_last_day, now_datetime, flt
from frappe import _
import traceback
import pandas as pd
import os
from datetime import datetime

def validate_unicommerce_settings():
    """Validates that all required Unicommerce Settings are properly configured."""
    settings = frappe.get_single("Unicommerce Settings")
    
    required_fields = {
        "tenant_url": "Tenant URL",
        "username": "Username", 
        "password": "Password",
        "facility": "Facility"
    }
    
    missing_fields = [label for field, label in required_fields.items() if not settings.get(field)]
    
    if missing_fields:
        frappe.throw(_("Missing required Unicommerce Settings: {0}").format(", ".join(missing_fields)))
    
    return settings

def create_export_job(export_job_type, export_columns, export_filters, facility, tenant_url, access_token):
    """Creates an export job in Unicommerce."""
    if not all([export_job_type, export_columns, export_filters, facility, tenant_url, access_token]):
        frappe.throw(_("Missing required parameters for export job creation"))
    
    export_url = f"{tenant_url.rstrip('/')}/services/rest/v1/export/job/create"
    
    payload = {
        "exportJobTypeName": export_job_type,
        "exportColums": export_columns,  # Note: API expects "exportColums" (typo in API)
        "exportFilters": export_filters,
        "frequency": "ONETIME"
    }
    
    headers = {
        'Accept': '*/*',
        'User-Agent': 'KhanalTech',
        'Content-Type': 'application/json',
        'facility': facility,  # facility is already a single facility name
        'Authorization': f'bearer {access_token}'
    }
    
    try:
        response = requests.post(export_url, headers=headers, data=json.dumps(payload), timeout=30)
        response.raise_for_status()
        result = response.json()
        
        if not result.get("successful", True):
            error_msg = result.get("message") or result.get("errors") or "Unknown error"
            frappe.log_error(f"Unicommerce export job creation failed: {error_msg}", "Unicommerce Export Error")
            raise Exception(f"Export job creation failed: {error_msg}")
        
        return result
        
    except Exception as e:
        frappe.log_error(f"Error creating export job: {str(e)}", "Unicommerce Export Error")
        raise

def get_export_columns_from_settings(settings):
    """Extracts and validates export columns from Unicommerce Settings."""
    columns_str = settings.get("export_columns") or settings.get("sale_order_export_columns") or ""
    
    if not columns_str:
        frappe.throw(_("Export columns not configured in Unicommerce Settings"))
    
    # Parse columns from JSON array or string
    try:
        if columns_str.strip().startswith('['):
            export_columns = json.loads(columns_str)
        elif '\n' in columns_str:
            export_columns = [col.strip() for col in columns_str.split('\n') if col.strip()]
        else:
            export_columns = [col.strip() for col in columns_str.split(',') if col.strip()]
    except json.JSONDecodeError:
        if '\n' in columns_str:
            export_columns = [col.strip() for col in columns_str.split('\n') if col.strip()]
        else:
            export_columns = [col.strip() for col in columns_str.split(',') if col.strip()]
    
    if not export_columns:
        frappe.throw(_("No valid export columns found in Unicommerce Settings"))
    
    return export_columns


@frappe.whitelist()
def trigger_previous_month_sale_order_export():
    """Creates Sale Orders export jobs for the previous full month using Unicommerce API for all facilities."""
    execution_start_time = now_datetime()
    
    try:
        settings = validate_unicommerce_settings()
        facility = settings.get("facility")
        if not facility:
            frappe.throw(_("Facility not configured in Unicommerce Settings"))
        
        export_columns = get_export_columns_from_settings(settings)
        
        
        uniware_session = AuthenticateUniware()
        if not uniware_session or not uniware_session.access_token:
            frappe.throw(_("Failed to authenticate with Unicommerce API"))

        # Calculate date range for the previous month
        previous_month_date = add_months(today(), -1)
        start_date = get_first_day(previous_month_date)
        end_date = get_last_day(previous_month_date)

        export_filters = [
            {
                "id": "channelCreatedDateRangeFilter",
                "dateRange": {
                    "start": start_date.strftime("%Y-%m-%d"),
                    "end": end_date.strftime("%Y-%m-%d")
                }
            }
        ]

        # Split facilities and create separate export jobs for each
        facilities = [f.strip() for f in facility.split(',') if f.strip()]
        job_results = []
        successful_jobs = 0
        failed_jobs = 0
        
        frappe.msgprint(_("Creating Sale Order export jobs for {0} to {1} for {2} facilities").format(
            start_date.strftime("%Y-%m-%d"), 
            end_date.strftime("%Y-%m-%d"),
            len(facilities)
        ))
        
        for single_facility in facilities:
            try:
                result = create_export_job(
                    export_job_type="Sale Orders",
                    export_columns=export_columns,
                    export_filters=export_filters,
                    facility=single_facility,
                    tenant_url=settings.tenant_url,
                    access_token=uniware_session.access_token
                )

                job_code = result.get("jobCode")
                if job_code:
                    successful_jobs += 1
                    job_results.append({
                        "facility": single_facility,
                        "job_code": job_code,
                        "status": "success"
                    })
                else:
                    failed_jobs += 1
                    error_details = result.get('errors') or result.get('message') or "Unknown error"
                    job_results.append({
                        "facility": single_facility,
                        "job_code": None,
                        "status": "failed",
                        "error": error_details
                    })
                    
            except Exception as e:
                failed_jobs += 1
                job_results.append({
                    "facility": single_facility,
                    "job_code": None,
                    "status": "failed",
                    "error": str(e)
                })
        
        execution_duration = flt((now_datetime() - execution_start_time).total_seconds(), 2)
        
        return {
            "success": True,
            "message": f"Created {successful_jobs} successful export jobs out of {len(facilities)} facilities",
            "date_range": f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
            "execution_duration": f"{execution_duration}s",
            "total_facilities": len(facilities),
            "successful_jobs": successful_jobs,
            "failed_jobs": failed_jobs,
            "job_details": job_results,
            "export_columns_count": len(export_columns)
        }
            
    except Exception as e:
        frappe.log_error(f"Export process failed: {str(e)}", "Unicommerce Export Error")
        frappe.throw(_("Export process failed: {0}").format(str(e)))

@frappe.whitelist()
def check_export_job_status(job_code):
    """Checks the status of an export job in Unicommerce."""
    if not job_code:
        frappe.throw(_("Job code is required"))
    
    try:
        settings = validate_unicommerce_settings()
        uniware_session = AuthenticateUniware()
        
        url = f"{settings.tenant_url}/services/rest/v1/export/job/status"
        payload = json.dumps({"jobCode": job_code})
        headers = {
            'Content-Type': 'application/json',
            'facility': 'kfl_ecom_warehouse',
            'Authorization': f'bearer {uniware_session.access_token}'
        }
        
        response = requests.post(url, headers=headers, data=payload)
        response.raise_for_status()
        
        return response.json()
        
    except Exception as e:
        frappe.log_error(f"Error checking export job status: {str(e)}", "Export Job Status Error")
        raise

@frappe.whitelist()
def download_and_merge_csv_files(job_codes_dict, max_wait_minutes=10):
    """Downloads CSV files and merges them into one Excel file with separate sheets."""
    import time
    
    try:
        settings = validate_unicommerce_settings()
        uniware_session = AuthenticateUniware()
        
        # Create output directory
        output_dir = frappe.get_site_path('private', 'files', 'unicommerce_exports')
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate filename with timestamp (including microseconds for uniqueness)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # Remove last 3 digits of microseconds
        excel_filename = f"unicommerce_export_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)
        
        # Ensure file doesn't already exist (extra safety)
        counter = 1
        while os.path.exists(excel_path):
            excel_filename = f"unicommerce_export_{timestamp}_{counter}.xlsx"
            excel_path = os.path.join(output_dir, excel_filename)
            counter += 1
        
        # Dictionary to store DataFrames and file info
        facility_dataframes = {}
        csv_files_info = {}
        pending_jobs = job_codes_dict.copy()
        
        frappe.msgprint(_("Waiting for export jobs to complete... (Max {0} minutes)").format(max_wait_minutes))
        
        # Wait for jobs to complete with timeout
        start_time = time.time()
        timeout_seconds = max_wait_minutes * 60
        
        while pending_jobs and (time.time() - start_time) < timeout_seconds:
            for facility, job_code in list(pending_jobs.items()):
                try:
                    # Check job status
                    status_result = check_export_job_status(job_code)
                    status = status_result.get("status")
                    
                    frappe.logger().info(f"Job {job_code} for {facility}: Status = {status}")
                    
                    if status == "COMPLETE" and status_result.get("filePath"):
                        csv_url = status_result["filePath"]
                        
                        # Extract CSV filename from URL
                        csv_filename = csv_url.split('/')[-1]
                        csv_local_path = os.path.join(output_dir, f"{facility}_{csv_filename}")
                        
                        # Download CSV data
                        csv_response = requests.get(csv_url)
                        csv_response.raise_for_status()
                        
                        # Save CSV file locally
                        with open(csv_local_path, 'wb') as f:
                            f.write(csv_response.content)
                        
                        # Read CSV into DataFrame
                        df = pd.read_csv(csv_local_path)
                        
                        # Store facility data (even if empty - ALL or NOTHING condition)
                        facility_dataframes[facility] = df
                        
                        csv_files_info[facility] = {
                            "csv_filename": f"{facility}_{csv_filename}",
                            "csv_path": csv_local_path,
                            "csv_url": csv_url,
                            "rows_count": len(df),
                            "job_code": job_code
                        }
                        
                        frappe.logger().info(f"Downloaded CSV for {facility}: {len(df)} rows")
                        
                        # Remove from pending
                        del pending_jobs[facility]
                        
                    elif status == "FAILED":
                        frappe.logger().error(f"❌ Job failed for {facility}: {status_result.get('errors', 'Unknown error')}")
                        del pending_jobs[facility]
                        
                    elif status in ["PROCESSING", "QUEUED", "PENDING"]:
                        frappe.logger().info(f"⏳ Job still processing for {facility}: {status}")
                        
                except Exception as e:
                    frappe.logger().error(f"Error processing {facility}: {str(e)}")
                    continue
            
            # Wait before checking again
            if pending_jobs:
                frappe.logger().info(f"Waiting for {len(pending_jobs)} jobs to complete...")
                time.sleep(30)  # Wait 30 seconds before checking again
        
        # Check if any jobs are still pending
        if pending_jobs:
            frappe.logger().warning(f"Timeout reached. {len(pending_jobs)} jobs still pending: {list(pending_jobs.keys())}")
        
        # ALL or NOTHING validation - ensure ALL facilities are included
        total_facilities = len(job_codes_dict)
        completed_facilities = len(facility_dataframes)
        
        if completed_facilities < total_facilities:
            missing_facilities = set(job_codes_dict.keys()) - set(facility_dataframes.keys())
            frappe.logger().error(f"Not all facilities completed. Expected: {total_facilities}, Completed: {completed_facilities}, Missing: {list(missing_facilities)}")
            frappe.throw(_(f"Not all facilities completed. Missing: {list(missing_facilities)}"))
        
        # Create Excel file with multiple sheets (ALL facilities)
        if facility_dataframes and len(facility_dataframes) == total_facilities:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for facility, df in facility_dataframes.items():
                    # Clean facility name for sheet name (Excel sheet name restrictions)
                    sheet_name = facility.replace('/', '_').replace('\\', '_')[:31]  # Max 31 chars
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Clean up individual CSV files after successful Excel creation
            deleted_csv_files = []
            for facility, file_info in csv_files_info.items():
                csv_path = file_info["csv_path"]
                try:
                    if os.path.exists(csv_path):
                        os.remove(csv_path)
                        deleted_csv_files.append(file_info["csv_filename"])
                        frappe.logger().info(f"Deleted CSV file: {file_info['csv_filename']}")
                except Exception as e:
                    frappe.logger().warning(f"Could not delete CSV file {file_info['csv_filename']}: {str(e)}")
            
            return {
                "success": True,
                "message": f"Created Excel file with ALL {len(facility_dataframes)} facilities",
                "excel_file": {
                    "filename": excel_filename,
                    "path": excel_path,
                    "absolute_path": os.path.abspath(excel_path)
                },
                "csv_files_processed": len(csv_files_info),
                "csv_files_deleted": deleted_csv_files,
                "output_directory": output_dir,
                "facilities_processed": sorted(list(facility_dataframes.keys())),
                "total_sheets": len(facility_dataframes),
                "all_facilities_included": len(facility_dataframes) == len(job_codes_dict),
                "facility_details": {
                    facility: {
                        "rows_count": info["rows_count"],
                        "job_code": info["job_code"],
                        "has_data": info["rows_count"] > 0
                    } for facility, info in csv_files_info.items()
                }
            }
        else:
            frappe.throw(_("No completed export jobs with data found. Check job status manually."))
            
    except Exception as e:
        frappe.log_error(f"Error downloading and merging CSV files: {str(e)}", "CSV Merge Error")
        raise

@frappe.whitelist()
def process_monthly_export_with_download():
    """Complete monthly export process: create jobs, wait for completion, and download merged Excel."""
    try:
        # Step 1: Create export jobs
        export_result = trigger_previous_month_sale_order_export()
        
        if not export_result.get("success"):
            frappe.throw(_("Failed to create export jobs"))
        
        # Step 2: Extract job codes
        job_codes_dict = {}
        job_details = export_result.get("job_details", [])
        
        if not job_details:
            frappe.throw(_("No job details found in export result"))
        
        for job_detail in job_details:
            if job_detail.get("status") == "success" and job_detail.get("job_code"):
                job_codes_dict[job_detail["facility"]] = job_detail["job_code"]
        
        if not job_codes_dict:
            frappe.throw(_("No export jobs were created successfully. All jobs failed with 'already scheduled/running' error. Please wait a few minutes and try again."))
        
        # Step 3: Download and merge CSV files
        merge_result = download_and_merge_csv_files(job_codes_dict)
        
        # Step 4: Create pivot analysis (NEW FEATURE)
        pivot_result = None
        if merge_result.get("success") and merge_result.get("excel_file", {}).get("path"):
            try:
                print("[MONTHLY EXPORT] Creating pivot analysis...")
                from khanal_tech_integrations.utils.Unicommerce_Automation.pivot_table_creator import create_complete_pivot_analysis
                
                excel_path = merge_result["excel_file"]["path"]
                print(f"[MONTHLY EXPORT] Excel file path: {excel_path}")
                
                # Check if file exists
                import os
                if not os.path.exists(excel_path):
                    print(f"[MONTHLY EXPORT] ❌ Excel file not found: {excel_path}")
                    pivot_result = {"status": "error", "error": "Excel file not found"}
                else:
                    print(f"[MONTHLY EXPORT] ✅ Excel file exists, creating pivot analysis...")
                    # Delay email until after validation + missing-order sync
                    pivot_result = create_complete_pivot_analysis(excel_path, email_notification=False)
                    
                    if pivot_result.get("status") == "success":
                        print("[MONTHLY EXPORT] ✅ Pivot analysis created successfully")
                        print(f"[MONTHLY EXPORT] Master data: {pivot_result.get('master_data', {}).get('total_rows', 0)} rows")
                        print(f"[MONTHLY EXPORT] Pivot tables: {len(pivot_result.get('pivot_tables', {}).get('created_tables', []))} created")
                    else:
                        print(f"[MONTHLY EXPORT] ❌ Pivot analysis failed: {pivot_result.get('error', 'Unknown error')}")
                        print(f"[MONTHLY EXPORT] Full error details: {pivot_result}")
                    
            except Exception as e:
                print(f"[MONTHLY EXPORT] ❌ Pivot analysis error: {str(e)}")
                import traceback
                traceback.print_exc()
                pivot_result = {"status": "error", "error": str(e)}
        else:
            print(f"[MONTHLY EXPORT] ⚠️ Cannot create pivot analysis - merge result: {merge_result}")
        
        # Step 5: Add validation remarks to the Excel file (automated)
        validation_result = None
        if merge_result.get("success") and merge_result.get("excel_file", {}).get("path"):
            try:
                print("[MONTHLY EXPORT] Adding validation remarks to Excel...")
                from khanal_tech_integrations.utils.Unicommerce_Automation.excel_validation_and_sync import add_validation_remarks_to_excel
                excel_path = merge_result["excel_file"]["path"]
                validation_result = add_validation_remarks_to_excel(excel_path)
                print(f"[MONTHLY EXPORT] Validation status: {validation_result.get('status')}")
            except Exception as e:
                print(f"[MONTHLY EXPORT] ❌ Validation remarks error: {str(e)}")
                validation_result = {"status": "error", "error": str(e)}

        # Step 6: Sync missing orders from Excel to DB (automated)
        missing_orders_sync_result = None
        if merge_result.get("success") and merge_result.get("excel_file", {}).get("path"):
            try:
                print("[MONTHLY EXPORT] Syncing missing orders from Excel to DB...")
                from khanal_tech_integrations.utils.Unicommerce_Automation.excel_validation_and_sync import sync_missing_orders_from_excel
                excel_path = merge_result["excel_file"]["path"]
                missing_orders_sync_result = sync_missing_orders_from_excel(excel_path)
                print(f"[MONTHLY EXPORT] Missing orders sync status: {missing_orders_sync_result.get('status')}")
            except Exception as e:
                print(f"[MONTHLY EXPORT] ❌ Missing orders sync error: {str(e)}")
                missing_orders_sync_result = {"status": "error", "error": str(e)}

        # Step 6.1: Reorder sheets: 1) Pivot Table 2) Master Data 3) kfl_ecom_warehouse 4) others 5) Validation Summary 6) Missing Orders
        try:
            import openpyxl
            excel_path = merge_result["excel_file"]["path"]
            wb = openpyxl.load_workbook(excel_path)
            desired_order = []
            # Ensure pivot first
            if 'Pivot Table' in wb.sheetnames:
                desired_order.append('Pivot Table')
            # Master next
            if 'Master Data' in wb.sheetnames:
                desired_order.append('Master Data')
            # Specific facility
            if 'kfl_ecom_warehouse' in wb.sheetnames:
                desired_order.append('kfl_ecom_warehouse')
            # Other facility/data sheets (exclude utility)
            utility = {'Pivot Table', 'Master Data', 'Validation Summary', 'Missing Orders'}
            others = [s for s in wb.sheetnames if s not in set(desired_order) and s not in utility]
            desired_order.extend(others)
            # Utility sheets last
            if 'Validation Summary' in wb.sheetnames:
                desired_order.append('Validation Summary')
            if 'Missing Orders' in wb.sheetnames:
                desired_order.append('Missing Orders')

            # Reorder by recreating workbook sheets
            # Note: openpyxl supports worksheet reordering via _sheets
            wb._sheets = [wb[s] for s in desired_order if s in wb.sheetnames]
            wb.save(excel_path)
        except Exception as e:
            print(f"[MONTHLY EXPORT] ⚠️ Sheet reorder skipped: {str(e)}")
        
        # Step 7: Email final Excel after all steps succeed
        final_email_result = None
        try:
            if pivot_result and pivot_result.get("status") == "success" and merge_result.get("success"):
                from khanal_tech_integrations.utils.Unicommerce_Automation.pivot_table_creator import _send_pivot_analysis_email
                _send_pivot_analysis_email(
                    merge_result["excel_file"]["path"],
                    pivot_result.get("master_data", {}),
                    pivot_result.get("pivot_table", {})
                )
                final_email_result = {"status": "sent"}
            else:
                final_email_result = {"status": "skipped", "reason": "pivot or merge not successful"}
        except Exception as e:
            final_email_result = {"status": "error", "error": str(e)}

        return {
            "export_jobs": export_result,
            "merged_file": merge_result,
            "pivot_analysis": pivot_result,
            "validation_remarks": validation_result,
            "missing_orders_sync": missing_orders_sync_result,
            "final_email": final_email_result
        }
        
    except Exception as e:
        frappe.log_error(f"Error in complete monthly export process: {str(e)}", "Monthly Export Process Error")
        raise
