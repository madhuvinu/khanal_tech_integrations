"""
Excel Report Generator for HANA DB Reporting System
Creates professional Excel reports with Sales Data and Pivot Analysis sheets
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import logging
from typing import Dict
from datetime import datetime
import os

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    """
    Excel report generator with Sales Data and Pivot Analysis sheets
    """
    
    def __init__(self):
        """
        Initialize Excel report generator
        """
        self.workbook = None
        self.worksheet = None
        
    def create_report(self, data: pd.DataFrame, report_config: Dict, filename: str = None) -> str:
        """
        Create Excel report with Sales Data and Pivot Analysis sheets
        
        Args:
            data: DataFrame containing the data
            report_config: Configuration for report formatting
            filename: Output filename (optional)
            
        Returns:
            str: Path to created Excel file
        """
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"hana_report_{timestamp}.xlsx"
            
            # Create workbook
            self.workbook = openpyxl.Workbook()
            
            # Remove default sheet
            self.workbook.remove(self.workbook.active)
            
            # Create main data sheet
            self._create_data_sheet(data, report_config)
            
            # Create pivot analysis sheet with both SALES INVOICE and CREDIT MEMO
            self._create_pivot_analysis_sheet(data, report_config)
            
            # Save workbook
            from .configuration import REPORT_SETTINGS
            output_dir = REPORT_SETTINGS.get("output_directory", "/tmp")
            os.makedirs(output_dir, exist_ok=True)
            filepath = os.path.join(output_dir, filename)
            self.workbook.save(filepath)
            
            logger.info(f"Excel report created successfully: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Failed to create Excel report: {str(e)}")
            raise
    
    def _create_data_sheet(self, data: pd.DataFrame, config: Dict):
        """
        Create main data sheet with formatting
        """
        try:
            # Create worksheet
            ws = self.workbook.create_sheet("Sales Data")
            
            # Add title
            title_cell = ws['A1']
            title_cell.value = config.get('title', 'SAP HANA Sales Analysis Report')
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Add timestamp
            timestamp_cell = ws['A2']
            timestamp_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            timestamp_cell.font = Font(size=10, italic=True)
            
            # Add data summary
            summary_cell = ws['A3']
            summary_cell.value = f"Total Records: {len(data):,} | Date Range: {data['Document Date'].min()} to {data['Document Date'].max()}"
            summary_cell.font = Font(size=10, italic=True)
            
            # Add data starting from row 5 (leave space for headers)
            start_row = 5
            
            # Clean and prepare data for Excel
            data_clean = data.copy()
            
            # Handle missing values
            data_clean = data_clean.fillna('')
            
            # Convert DataFrame to worksheet starting from the correct row
            for r_idx, row_data in enumerate(dataframe_to_rows(data_clean, index=False, header=True)):
                current_row = start_row + r_idx
                for c_idx, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=c_idx, value=value)
            
            # Apply header formatting
            header_row = start_row
            for col_num, column_title in enumerate(data.columns, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply column-specific formatting
            self._apply_column_formatting(ws, data, start_row + 1)
            
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply data formatting
            for row in range(start_row, start_row + len(data) + 1):
                for col in range(1, len(data.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Create table
            table = Table(displayName="SalesData", ref=f"A{start_row}:{get_column_letter(len(data.columns))}{start_row + len(data)}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(table)
            
            logger.info("Data sheet created successfully")
            
        except Exception as e:
            logger.error(f"Failed to create data sheet: {str(e)}")
            raise
    
    def _apply_column_formatting(self, ws, data: pd.DataFrame, start_row: int):
        """
        Apply column-specific formatting for better readability
        
        Args:
            ws: Worksheet object
            data: DataFrame with data
            start_row: Starting row for data
        """
        try:
            # Define formatting rules
            currency_columns = ['MRP', 'Unit Price', 'INR Price', 'Taxable Value', 'Total_Disc', 
                              'Net Taxable Value', 'CGST Tax Amount', 'SGST Tax Amount', 
                              'IGST Tax Amount', 'Freight Charges', 'Insurance', 'Total Value',
                              'COGS per unit', 'COGS', 'Gross Margin']
            
            percentage_columns = ['Diff %', 'GST_Rate', 'Gross Margin %']
            
            integer_columns = ['DocNum', 'LineNum', 'Quantity', 'MIS Weight']
            
            date_columns = ['Document Date', 'Order Date']
            
            # Apply formatting to each column
            for col_idx, col_name in enumerate(data.columns, 1):
                if col_name in currency_columns:
                    # Currency formatting
                    for row in range(start_row, start_row + len(data)):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value and pd.notna(cell.value) and cell.value != '':
                            cell.number_format = '#,##0.00'
                
                elif col_name in percentage_columns:
                    # Percentage formatting
                    for row in range(start_row, start_row + len(data)):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value and pd.notna(cell.value) and cell.value != '':
                            cell.number_format = '0.00%'
                
                elif col_name in integer_columns:
                    # Integer formatting
                    for row in range(start_row, start_row + len(data)):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value and pd.notna(cell.value) and cell.value != '':
                            cell.number_format = '#,##0'
                
                elif col_name in date_columns:
                    # Date formatting
                    for row in range(start_row, start_row + len(data)):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value and pd.notna(cell.value) and cell.value != '':
                            cell.number_format = 'dd-mmm-yyyy'
                
                # Auto-adjust column width
                column_letter = get_column_letter(col_idx)
                max_length = max(
                    len(str(col_name)),
                    max(len(str(cell.value)) for cell in ws[column_letter] if cell.value)
                )
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            logger.info("Column formatting applied successfully")
            
        except Exception as e:
            logger.error(f"Failed to apply column formatting: {str(e)}")
            raise
    
    def _create_pivot_analysis_sheet(self, data: pd.DataFrame, config: Dict):
        """
        Create pivot analysis sheet with both SALES INVOICE and CREDIT MEMO data
        """
        try:
            # Create worksheet
            ws = self.workbook.create_sheet("Pivot Analysis")
            
            # Add title
            title_cell = ws['A1']
            title_cell.value = "Sales Analysis - Customer-wise Tax Summary"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Add timestamp
            timestamp_cell = ws['A2']
            timestamp_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            timestamp_cell.font = Font(size=10, italic=True)
            
            # Prepare data for pivot table
            pivot_data = data.copy()
            
            # Process both SALES INVOICE and CREDIT MEMO
            document_types = ['SALES INVOICE', 'CREDIT MEMO']
            current_row = 4
            
            for doc_type in document_types:
                # Filter data for current document type
                if 'Document' in pivot_data.columns:
                    filtered_data = pivot_data[pivot_data['Document'] == doc_type]
                else:
                    continue
                
                if filtered_data.empty:
                    continue
                
                # Add filter section (Document filter)
                ws.cell(row=current_row, column=1, value="Document").font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=doc_type).font = Font(bold=True)
                
                # Add column headers
                header_row = current_row + 2
                headers = ["Row Labels", "Sum of Taxable Value", "Sum of CGST Tax Amount", 
                          "Sum of SGST Tax Amount", "Sum of IGST Tax Amount"]
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Group data by Customer Name and calculate sums
                if 'Customer Name' in filtered_data.columns:
                    # Convert numeric columns to proper types for aggregation
                    numeric_cols = ['Taxable Value', 'CGST Tax Amount', 'SGST Tax Amount', 'IGST Tax Amount']
                    for col in numeric_cols:
                        if col in filtered_data.columns:
                            filtered_data[col] = pd.to_numeric(filtered_data[col], errors='coerce')
                    
                    # Group by Customer Name and sum the values
                    grouped_data = filtered_data.groupby('Customer Name')[numeric_cols].sum().reset_index()
                    
                    # Sort by Taxable Value descending
                    grouped_data = grouped_data.sort_values('Taxable Value', ascending=False)
                    
                    # Add data rows
                    data_start_row = header_row + 1
                    for idx, (_, row) in enumerate(grouped_data.iterrows()):
                        current_data_row = data_start_row + idx
                        
                        # Customer Name
                        ws.cell(row=current_data_row, column=1, value=row['Customer Name'])
                        
                        # Taxable Value
                        ws.cell(row=current_data_row, column=2, value=row['Taxable Value'])
                        
                        # CGST Tax Amount
                        ws.cell(row=current_data_row, column=3, value=row['CGST Tax Amount'])
                        
                        # SGST Tax Amount
                        ws.cell(row=current_data_row, column=4, value=row['SGST Tax Amount'])
                        
                        # IGST Tax Amount
                        ws.cell(row=current_data_row, column=5, value=row['IGST Tax Amount'])
                    
                    # Add Grand Total row
                    total_row = data_start_row + len(grouped_data)
                    ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)
                    
                    # Calculate totals
                    total_taxable = grouped_data['Taxable Value'].sum()
                    total_cgst = grouped_data['CGST Tax Amount'].sum()
                    total_sgst = grouped_data['SGST Tax Amount'].sum()
                    total_igst = grouped_data['IGST Tax Amount'].sum()
                    
                    ws.cell(row=total_row, column=2, value=total_taxable).font = Font(bold=True)
                    ws.cell(row=total_row, column=3, value=total_cgst).font = Font(bold=True)
                    ws.cell(row=total_row, column=4, value=total_sgst).font = Font(bold=True)
                    ws.cell(row=total_row, column=5, value=total_igst).font = Font(bold=True)
                    
                    # Apply formatting to all data cells for this section
                    for row_idx in range(header_row, total_row + 1):
                        for col_idx in range(1, 6):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            
                            # Add border
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            cell.border = thin_border
                            
                            # Format numeric columns (columns 2-5)
                            if col_idx >= 2 and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                            
                            # Center align all cells
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Update current row for next section
                    current_row = total_row + 3  # Add spacing between sections
                else:
                    current_row = header_row + 3
            
            # Auto-adjust column widths
            column_widths = [50, 20, 20, 20, 20]  # Adjust based on content
            for col_idx, width in enumerate(column_widths, 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = width
            
            logger.info("Pivot analysis sheet created successfully with both SALES INVOICE and CREDIT MEMO")
            
        except Exception as e:
            logger.error(f"Failed to create pivot analysis sheet: {str(e)}")
            raise

# Report configuration templates
REPORT_CONFIGS = {
    "comprehensive_sales": {
        "title": "SAP HANA Comprehensive Sales Analysis Report",
        "include_summary": False,
        "include_pivot_tables": False,
        "include_charts": False,
        "formatting": {
            "header_style": {"bg_color": "366092", "font_color": "white", "bold": True},
            "data_style": {"border": True, "alignment": "center"},
            "number_format": {"currency": True, "decimal_places": 2}
        }
    }
}

def create_comprehensive_sales_report(data: pd.DataFrame, filename: str = None) -> str:
    """
    Create comprehensive sales report with Sales Data and Pivot Analysis sheets
    
    Args:
        data: Sales DataFrame
        filename: Output filename
        
    Returns:
        str: Path to created Excel file
    """
    try:
        generator = ExcelReportGenerator()
        config = REPORT_CONFIGS["comprehensive_sales"]
        return generator.create_report(data, config, filename)
        
    except Exception as e:
        logger.error(f"Failed to create comprehensive sales report: {str(e)}")
        raise

if __name__ == "__main__":
    # Test the Excel generator
    try:
        # Create sample data
        sample_data = pd.DataFrame({
            'Company': ['KFPL-KA', 'KFPL-MH', 'KFPL-TN'],
            'Customer Name': ['Customer A', 'Customer B', 'Customer C'],
            'Total Value': [100000, 150000, 200000],
            'Channel': ['E-com', 'Offline HN', 'Export'],
            'Month': ['Jan-25', 'Jan-25', 'Jan-25'],
            'Document': ['SALES INVOICE', 'SALES INVOICE', 'CREDIT MEMO'],
            'Taxable Value': [90000, 135000, 180000],
            'CGST Tax Amount': [8100, 12150, 16200],
            'SGST Tax Amount': [8100, 12150, 16200],
            'IGST Tax Amount': [0, 0, 0],
            'Document Date': pd.to_datetime(['2025-01-15', '2025-01-16', '2025-01-17'])
        })
        
        # Create report
        generator = ExcelReportGenerator()
        config = REPORT_CONFIGS["comprehensive_sales"]
        filepath = generator.create_report(sample_data, config)
        
        print(f"✅ Excel report created successfully: {filepath}")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")