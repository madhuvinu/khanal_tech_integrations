"""
Excel Validation and Missing Orders Sync
======================================

Combined module for:
1. Adding validation remarks to Excel files
2. Syncing missing orders from Excel to database
"""

import frappe
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import shutil
from datetime import datetime
from typing import Dict

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _safe_convert_to_float(value) -> float:
    """Safely convert value to float"""
    try:
        if value is None or pd.isna(value) or value == '':
            return 0.0
        return float(str(value).strip())
    except:
        return 0.0

def _normalize_header(name: str) -> str:
    """Normalize header for matching (case/space/underscore insensitive)."""
    if not isinstance(name, str):
        return ""
    return ''.join(str(name).strip().lower().replace('_', ' ').split())

def _find_best_column(df: pd.DataFrame, aliases: list) -> str:
    """Find the best matching column in df given a list of alias names."""
    if df is None or df.empty or not isinstance(df.columns, pd.Index):
        return None
    normalized_cols = { _normalize_header(col): col for col in df.columns }
    for alias in aliases:
        norm = _normalize_header(alias)
        if norm in normalized_cols:
            return normalized_cols[norm]
    return None

def _norm_id(value) -> str:
    """Normalize an order id string for comparison."""
    if value is None:
        return ''
    return str(value).strip().lower()

def _sanitize_order_code_for_api(value: str) -> str:
    """Remove leading non-alphanumeric symbols (e.g., Excel leading apostrophe, backticks) before API calls."""
    if value is None:
        return ''
    s = str(value)
    # strip whitespace and zero-width/invisible chars
    s = s.strip().replace('\u200b', '').replace('\ufeff', '')
    # remove any leading characters until first [A-Za-z0-9_-] (keep hyphen/underscore)
    # This will remove backticks, apostrophes, and other special characters
    return re.sub(r"^[^A-Za-z0-9_-]+", "", s)

# =============================================================================
# VALIDATION REMARKS FUNCTION
# =============================================================================

@frappe.whitelist()
def add_validation_remarks_to_excel(excel_file_path: str) -> dict:
    """
    Add validation columns to existing Excel file
    
    Args:
        excel_file_path: Path to the Excel file
        
    Returns:
        Dictionary with results
    """
    try:
        print(f"[VALIDATION] Starting: {excel_file_path}")
        
        # Step 1: Load database data
        print("[VALIDATION] Loading DB data...")
        
        # Get orders from database
        orders = frappe.db.sql("""
            SELECT uniware_id, channel_id, status
            FROM `tabUnicommerce Orders`
        """, as_dict=True)
        
        # Get items from database
        items = frappe.db.sql("""
            SELECT oli.parent, oli.itemsku, oli.selling_price, oli.total_price, oli.statusCode
            FROM `tabOrder Line Items` oli
            WHERE oli.parenttype = 'Unicommerce Orders'
        """, as_dict=True)
        
        # Create lookup dictionaries
        orders_dict = {order['channel_id']: order for order in orders}
        # Add unified lookup by either channel_id or uniware_id
        orders_by_any = {}
        # Create a fast lookup dictionary: uniware_id -> order (for O(1) access)
        orders_by_uniware_id = {}
        for order in orders:
            if order.get('channel_id'):
                orders_by_any[str(order['channel_id'])] = order
            if order.get('uniware_id'):
                orders_by_any[str(order['uniware_id'])] = order
                # Create fast lookup for uniware_id -> order
                orders_by_uniware_id[str(order['uniware_id'])] = order
        
        # OPTIMIZED: Use dictionary lookup instead of nested loop
        # This reduces complexity from O(n*m) to O(n) where n=items, m=orders
        items_dict = {}
        for item in items:
            # Use O(1) dictionary lookup instead of O(m) loop
            order = orders_by_uniware_id.get(str(item['parent']))
            if order and order.get('channel_id'):
                key = f"{order['channel_id']}_{item['itemsku']}"
                items_dict[key] = item
        
        print(f"[VALIDATION] DB orders: {len(orders)} | keys: {len(orders_by_any)} | items: {len(items_dict)}")
        
        # Step 2: Create backup
        backup_path = excel_file_path.replace('.xlsx', '_backup.xlsx')
        shutil.copy2(excel_file_path, backup_path)
        print(f"[VALIDATION] Backup: {backup_path}")
        
        # Step 3: Process Excel file
        workbook = openpyxl.load_workbook(excel_file_path)
        
        total_orders = 0
        orders_found = 0
        total_items = 0
        items_found = 0
        missing_orders_accumulator = []
        unique_orders_processed = set()  # Track unique orders to avoid double counting
        
        for sheet_name in workbook.sheetnames:
            # Skip non-data sheets explicitly
            if str(sheet_name).strip().lower() in { 'pivot table', 'validation summary', 'missing orders' }:
                print(f"[VALIDATION] Skipped {sheet_name}: utility sheet")
                continue
            print(f"[VALIDATION] Sheet: {sheet_name}")
            
            try:
                # Read sheet data
                df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
                if len(df) > 0:
                    # Flexible header detection
                    order_code_col = _find_best_column(df, [
                        'Sale Order Code', 'Sales Order Code', 'SaleOrderCode', 'SalesOrderCode',
                        'Display Order Code', 'Order Code', 'DisplayOrderCode', 'OrderCode',
                        'Channel Order Code', 'Channel ID', 'Channel Id', 'ChannelId'
                    ])
                    item_sku_col = _find_best_column(df, [
                        'Item SKU Code', 'SKU', 'Item SKU', 'Sku Code', 'Item Code', 'ItemSKU'
                    ])
                    selling_price_col = _find_best_column(df, [
                        'Selling Price', 'SellingPrice', 'Unit Selling Price', 'Unit Price', 'Price'
                    ])

                    if not order_code_col:
                        print(f"[VALIDATION] Skipped {sheet_name}: no order code column")
                        continue
                    if not item_sku_col:
                        pass
                    if not selling_price_col:
                        pass
                    
                    # Add validation columns
                    db_status_list = []
                    validation_remarks_list = []
                    remarks_list = []
                    
                    for _, row in df.iterrows():
                        display_order_code = row.get(order_code_col, '') if order_code_col else ''
                        item_sku = row.get(item_sku_col, '') if item_sku_col else ''
                        
                        # Check order (by uniware_id or channel_id) - only count unique orders
                        if not order_code_col:
                            order_status = '❌ No Order Code Column'
                        else:
                            # Only count this order once, even if it has multiple items
                            if display_order_code and display_order_code not in unique_orders_processed:
                                unique_orders_processed.add(display_order_code)
                                total_orders += 1
                                
                                db_order = orders_by_any.get(str(display_order_code))
                                if db_order:
                                    order_status = '✅ Found in DB'
                                    orders_found += 1
                                else:
                                    order_status = '❌ Missing in DB'
                            else:
                                # This is a duplicate order (multiple items), use previous status
                                db_order = orders_by_any.get(str(display_order_code))
                                if db_order:
                                    order_status = '✅ Found in DB'
                                else:
                                    order_status = '❌ Missing in DB'
                        
                        # Check item
                        if order_code_col and item_sku_col and display_order_code and item_sku:
                            key = f"{display_order_code}_{item_sku}"
                            if key in items_dict:
                                item_status = '✅ Found in DB'
                                items_found += 1
                                
                                # Check price match
                                db_item = items_dict[key]
                                excel_selling_price = _safe_convert_to_float(row.get(selling_price_col)) if selling_price_col else None
                                db_selling_price = _safe_convert_to_float(db_item.get('selling_price'))
                                
                                if excel_selling_price is not None and excel_selling_price != db_selling_price:
                                    item_status += f' (Price mismatch: Excel={excel_selling_price} vs DB={db_selling_price})'
                            else:
                                item_status = '❌ Missing in DB'
                        else:
                            if not item_sku_col:
                                item_status = '❌ No SKU Column'
                            else:
                                item_status = '❌ No SKU'
                        
                        total_items += 1
                        
                        # Combine remarks
                        db_status_list.append(order_status)
                        validation_remarks_list.append(f"Order: {order_status} | Item: {item_status}")
                        # Build single 'Remarks' per row
                        excel_status_col = excel_status_col if 'excel_status_col' in locals() else None
                        excel_status = str(row.get(excel_status_col, '')).strip() if excel_status_col else ''
                        if not order_code_col:
                            remarks = '❌ No Order Code Column'
                        else:
                            db_order = orders_by_any.get(str(display_order_code))
                            if db_order:
                                db_status_value = str(db_order.get('status') or '').strip()
                                if excel_status and db_status_value and excel_status != db_status_value:
                                    remarks = f"✅ Present | Status MISMATCH (Excel={excel_status} vs DB={db_status_value})"
                                else:
                                    remarks = '✅ Present | Status match'
                            else:
                                remarks = '❌ Missing in DB'
                                # Only add to missing orders if not already added
                                if display_order_code and not any(entry['order_code'] == str(display_order_code) for entry in missing_orders_accumulator):
                                    missing_orders_accumulator.append({'order_code': str(display_order_code), 'excel_status': excel_status})
                        remarks_list.append(remarks)
                    
                    # Add new columns
                    df['DB_Status'] = db_status_list
                    df['Validation_Remarks'] = validation_remarks_list
                    df['Remarks'] = remarks_list
                    
                    # Clear and rewrite sheet
                    worksheet = workbook[sheet_name]
                    worksheet.delete_rows(1, worksheet.max_row)
                    
                    # Add headers first
                    for col_num, column_title in enumerate(df.columns, 1):
                        worksheet.cell(row=1, column=col_num, value=column_title)
                        worksheet.cell(row=1, column=col_num).font = Font(bold=True)
                    
                    # Add data with new columns (convert numpy array to list)
                    for row_idx, row_data in enumerate(df.values, start=2):
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    print(f"[VALIDATION] Updated: {sheet_name}")
                else:
                    print(f"[VALIDATION] Skipped {sheet_name}: empty")
            except Exception as e:
                print(f"[VALIDATION] Error {sheet_name}: {str(e)}")
                continue
        
        # Step 4a: Create Missing Orders sheet and repull
        # Deduplicate and repull
        unique_missing = []
        seen = set()
        for entry in missing_orders_accumulator:
            code = entry.get('order_code')
            if code and code not in seen:
                seen.add(code)
                unique_missing.append(entry)

        pull_results = []
        if unique_missing:
            try:
                from .unicommerceFile.unicommerce_clean import get_single_order, push_new_orders
            except Exception as e:
                print(f"[VALIDATION] Could not import repull functions: {str(e)}")
                get_single_order = None
                push_new_orders = None

            for entry in unique_missing:
                order_code = entry.get('order_code')
                excel_status = entry.get('excel_status')
                pull_status = 'Skipped'
                notes = ''
                if get_single_order and push_new_orders:
                    try:
                        # First try with original code
                        data = get_single_order(order_code)
                        if not data:
                            # Retry with sanitized code (remove leading symbol)
                            sanitized = _sanitize_order_code_for_api(order_code)
                            if sanitized != order_code:
                                data = get_single_order(sanitized)
                                if data:
                                    notes = f"Retried with sanitized code: {sanitized}"
                        if data:
                            saved = push_new_orders(data, update=False)
                            pull_status = 'Success' if saved else 'Failed'
                            if not notes:
                                notes = '' if saved else 'push_new_orders returned False'
                        else:
                            pull_status = 'Failed'
                            notes = 'get_single_order returned no data'
                    except Exception as e:
                        pull_status = 'Failed'
                        notes = str(e)
                else:
                    notes = 'Repull functions unavailable'
                pull_results.append({'order_code': order_code, 'excel_status': excel_status, 'pull_status': pull_status, 'notes': notes})

        # Remove existing and create Missing Orders sheet at position 0
        if 'Missing Orders' in workbook.sheetnames:
            workbook.remove(workbook['Missing Orders'])
        missing_ws = workbook.create_sheet('Missing Orders', 0)
        headers = ['Order Code', 'Excel Status', 'Pull Status', 'Notes']
        for col_idx, title in enumerate(headers, 1):
            missing_ws.cell(row=1, column=col_idx, value=title)
            missing_ws.cell(row=1, column=col_idx).font = Font(bold=True)
        for row_idx, result in enumerate(pull_results, start=2):
            missing_ws.cell(row=row_idx, column=1, value=result.get('order_code'))
            missing_ws.cell(row=row_idx, column=2, value=result.get('excel_status'))
            missing_ws.cell(row=row_idx, column=3, value=result.get('pull_status'))
            missing_ws.cell(row=row_idx, column=4, value=result.get('notes'))

        # Step 4b: Add summary sheet after Missing Orders
        summary_sheet = workbook.create_sheet("Validation Summary", 1)
        summary_sheet['A1'] = "Unicommerce Data Validation Summary"
        summary_sheet['A1'].font = Font(bold=True, size=16)
        
        summary_data = [
            ["Validation Date:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Total Unique Orders Processed:", total_orders],
            ["Orders Found in Database:", orders_found],
            ["Orders Missing in Database:", total_orders - orders_found],
            ["Total Items Processed:", total_items],
            ["Items Found in Database:", items_found],
            ["Items Missing in Database:", total_items - items_found],
            ["Missing Orders Identified:", len(unique_missing)],
            ["Missing Orders Repull Attempted:", len(pull_results)],
            ["Repull Success Count:", sum(1 for r in pull_results if r.get('pull_status') == 'Success')],
            ["Repull Failed Count:", sum(1 for r in pull_results if r.get('pull_status') == 'Failed')]
        ]
        
        for i, (label, value) in enumerate(summary_data, 2):
            summary_sheet[f'A{i}'] = label
            summary_sheet[f'B{i}'] = value
            if label.endswith(':'):
                summary_sheet[f'A{i}'].font = Font(bold=True)
        
        # Step 5: Save modified file
        workbook.save(excel_file_path)
        workbook.close()
        
        print(f"[VALIDATION] Saved: {excel_file_path}")
        
        # Calculate percentages
        order_match_percentage = (orders_found/total_orders*100) if total_orders > 0 else 0
        item_match_percentage = (items_found/total_items*100) if total_items > 0 else 0
        
        return {
            'status': 'success',
            'modified_excel_path': excel_file_path,
            'summary': {
                'total_orders_processed': total_orders,
                'orders_found_in_db': orders_found,
                'orders_missing_in_db': total_orders - orders_found,
                'order_match_percentage': order_match_percentage,
                'total_items_processed': total_items,
                'items_found_in_db': items_found,
                'items_missing_in_db': total_items - items_found,
                'item_match_percentage': item_match_percentage
            },
            'message': 'Validation remarks successfully added to Excel file'
        }
        
    except Exception as e:
        error_msg = f"Error adding validation remarks: {str(e)}"
        print(f"[VALIDATION] ERROR: {error_msg}")
        
        return {
            'status': 'error',
            'error': error_msg
        }

# =============================================================================
# MISSING ORDERS SYNC FUNCTION
# =============================================================================

@frappe.whitelist()
def sync_missing_orders_from_excel(excel_file_path: str) -> Dict:
    """
    Simple function to sync missing orders from Excel to database
    
    Args:
        excel_file_path: Path to the Excel file
        
    Returns:
        Dictionary with sync results
    """
    try:
        print(f"[MISSING ORDERS SYNC] Starting sync for: {excel_file_path}")
        
        # Step 1: Get all order codes from database
        db_orders = frappe.db.sql("""
            SELECT channel_id FROM `tabUnicommerce Orders` 
            WHERE channel_id IS NOT NULL AND channel_id != ''
        """, as_dict=True)
        
        db_order_codes = set(order['channel_id'] for order in db_orders)
        print(f"[MISSING ORDERS SYNC] Found {len(db_order_codes)} orders in database")
        
        # Step 2: Get all order codes from Excel
        excel_order_codes = set()
        
        # Read all sheets
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
        for sheet_name in workbook.sheetnames:
            try:
                df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
                if 'Sale Order Code' in df.columns:
                    sheet_codes = set(df['Sale Order Code'].dropna().astype(str))
                    excel_order_codes.update(sheet_codes)
                    print(f"[MISSING ORDERS SYNC] Sheet {sheet_name}: {len(sheet_codes)} orders")
            except Exception as e:
                print(f"[MISSING ORDERS SYNC] Error reading sheet {sheet_name}: {str(e)}")
        
        workbook.close()
        
        # Step 3: Find missing orders
        missing_orders = excel_order_codes - db_order_codes
        print(f"[MISSING ORDERS SYNC] Missing orders: {len(missing_orders)}")
        
        if not missing_orders:
            return {
                'status': 'success',
                'message': 'No missing orders found - database is up to date',
                'missing_count': 0
            }
        
        # Step 4: Fetch missing orders using existing function
        from .unicommerceFile.unicommerce_clean import get_single_order, push_new_orders
        
        success_count = 0
        failed_count = 0
        
        for order_code in list(missing_orders)[:50]:  # Limit to first 50 for safety
            try:
                print(f"[MISSING ORDERS SYNC] Fetching order: {order_code}")
                
                # First try with original code
                order_data = get_single_order(order_code)
                
                if not order_data:
                    # Retry with sanitized code (remove leading special characters)
                    sanitized_code = _sanitize_order_code_for_api(order_code)
                    if sanitized_code != order_code and sanitized_code:
                        print(f"[MISSING ORDERS SYNC] Retrying with sanitized code: {sanitized_code}")
                        order_data = get_single_order(sanitized_code)
                
                if order_data:
                    # Use existing push_new_orders function
                    result = push_new_orders(order_data, update=False)
                    if result:
                        success_count += 1
                        print(f"[MISSING ORDERS SYNC] ✅ Success: {order_code}")
                    else:
                        failed_count += 1
                        print(f"[MISSING ORDERS SYNC] ❌ Failed to save: {order_code}")
                else:
                    failed_count += 1
                    print(f"[MISSING ORDERS SYNC] ❌ Not found: {order_code}")
                    
            except Exception as e:
                failed_count += 1
                print(f"[MISSING ORDERS SYNC] ❌ Error: {order_code} - {str(e)}")
        
        return {
            'status': 'success',
            'message': f'Sync completed. Success: {success_count}, Failed: {failed_count}',
            'missing_count': len(missing_orders),
            'success_count': success_count,
            'failed_count': failed_count
        }
        
    except Exception as e:
        error_msg = f"Sync failed: {str(e)}"
        print(f"[MISSING ORDERS SYNC] ERROR: {error_msg}")
        
        return {
            'status': 'error',
            'error': error_msg
        }
